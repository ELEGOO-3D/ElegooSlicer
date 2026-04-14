#!/usr/bin/env python3
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, Tuple

import polib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


@dataclass(frozen=True)
class EntryKey:
    msgctxt: str
    msgid: str


@dataclass(frozen=True)
class EntryValue:
    msgid_plural: str
    translation: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export zh_CN/en/de translations and keys from PO files to Excel."
    )
    parser.add_argument(
        "--i18n-dir",
        type=Path,
        default=Path("localization/i18n"),
        help="Path to i18n directory (default: localization/i18n)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("localization/i18n/zh_en_de_translations.xlsx"),
        help="Output xlsx path (default: localization/i18n/zh_en_de_translations.xlsx)",
    )
    parser.add_argument(
        "--sheet-name",
        default="zh_en_de",
        help="Excel sheet name (default: zh_en_de)",
    )
    parser.add_argument(
        "--include-commented",
        action="store_true",
        help="Include fuzzy entries. Obsolete (#~) entries are always excluded.",
    )
    return parser.parse_args()


def po_file_path(i18n_dir: Path, locale: str) -> Path:
    return i18n_dir / locale / f"ElegooSlicer_{locale}.po"


def should_export_entry(entry: polib.POEntry, include_commented: bool) -> bool:
    if entry.obsolete or entry.msgid == "":
        return False
    if include_commented:
        return True

    # By default skip fuzzy entries in review state.
    if "fuzzy" in entry.flags:
        return False
    return True


def render_translation(entry: polib.POEntry) -> str:
    if entry.msgstr_plural:
        parts = []
        for idx in sorted(entry.msgstr_plural.keys(), key=lambda x: int(x)):
            parts.append(f"[{idx}] {entry.msgstr_plural[idx]}")
        return "\n".join(parts)
    return entry.msgstr or ""


def load_po_entries(path: Path, include_commented: bool) -> Dict[EntryKey, EntryValue]:
    po = polib.pofile(str(path), encoding="utf-8")
    data: Dict[EntryKey, EntryValue] = {}

    for entry in po:
        if not should_export_entry(entry, include_commented):
            continue

        key = EntryKey(msgctxt=entry.msgctxt or "", msgid=entry.msgid)
        data[key] = EntryValue(
            msgid_plural=entry.msgid_plural or "",
            translation=render_translation(entry),
        )

    return data


def merged_keys(*maps: Dict[EntryKey, EntryValue]) -> Iterable[EntryKey]:
    all_keys = set()
    for m in maps:
        all_keys.update(m.keys())
    return sorted(all_keys, key=lambda k: (k.msgctxt, k.msgid))


def build_workbook(
    rows: Iterable[Tuple[EntryKey, str, str, str, str]],
    output_path: Path,
    sheet_name: str,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    headers = ["key", "msgctxt", "msgid", "msgid_plural", "zh_CN", "en", "de"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    for col, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill

    for row_index, (entry_key, msgid_plural, zh_cn, en, de) in enumerate(rows, start=2):
        key_text = f"{entry_key.msgctxt}|{entry_key.msgid}" if entry_key.msgctxt else entry_key.msgid
        sheet.cell(row=row_index, column=1, value=key_text)
        sheet.cell(row=row_index, column=2, value=entry_key.msgctxt)
        sheet.cell(row=row_index, column=3, value=entry_key.msgid)
        sheet.cell(row=row_index, column=4, value=msgid_plural)
        sheet.cell(row=row_index, column=5, value=zh_cn)
        sheet.cell(row=row_index, column=6, value=en)
        sheet.cell(row=row_index, column=7, value=de)

    sheet.column_dimensions["A"].width = 90
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 90
    sheet.column_dimensions["D"].width = 90
    sheet.column_dimensions["E"].width = 60
    sheet.column_dimensions["F"].width = 60
    sheet.column_dimensions["G"].width = 60

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    i18n_dir: Path = args.i18n_dir

    zh_path = po_file_path(i18n_dir, "zh_CN")
    en_path = po_file_path(i18n_dir, "en")
    de_path = po_file_path(i18n_dir, "de")

    missing_files = [p for p in [zh_path, en_path, de_path] if not p.exists()]
    if missing_files:
        missing_text = "\n".join(str(p) for p in missing_files)
        raise FileNotFoundError(f"Required PO files not found:\n{missing_text}")

    zh_map = load_po_entries(zh_path, include_commented=args.include_commented)
    en_map = load_po_entries(en_path, include_commented=args.include_commented)
    de_map = load_po_entries(de_path, include_commented=args.include_commented)

    export_rows = []
    for key in merged_keys(zh_map, en_map, de_map):
        msgid_plural = ""
        if key in zh_map:
            msgid_plural = zh_map[key].msgid_plural
        elif key in en_map:
            msgid_plural = en_map[key].msgid_plural
        elif key in de_map:
            msgid_plural = de_map[key].msgid_plural

        export_rows.append(
            (
                key,
                msgid_plural,
                zh_map[key].translation if key in zh_map else "",
                en_map[key].translation if key in en_map else "",
                de_map[key].translation if key in de_map else "",
            )
        )

    build_workbook(export_rows, args.output, args.sheet_name)
    print(f"Exported {len(export_rows)} rows to: {args.output}")


if __name__ == "__main__":
    main()
